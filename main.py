import openpyxl , sys
from routePlanner import *
from tubeMap import TubeMap
from datetime import datetime, timedelta
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QMainWindow



class Node(object):

    def __init__(self, distance=None, next_node=None, next_node_reference=None, prev_node=None, prev_node_reference=None, trainLine=None):
        self.distance = distance
        self.next_node = next_node
        self.next_node_reference = next_node_reference
        self.prev_node = prev_node
        self.prev_node_reference = prev_node_reference
        self.trainLine = trainLine

class DoubleLinkedList(object):

    def __init__(self, head=None, end=None):
        self.head = head
        self.end = end

    def traverse(self):
        curr_node = self.head
        while curr_node != None:
            #print(curr_node.prev_node, curr_node.next_node, curr_node.distance)
            curr_node = curr_node.next_node_reference
        
    def addToGraph(self,graph):
        curr_node = self.head
        while curr_node != None:
            if curr_node.prev_node in graph:
                graph[curr_node.prev_node][curr_node.next_node] = [int(float(curr_node.distance))]
            else:
                graph[curr_node.prev_node] = {}
                graph[curr_node.prev_node][curr_node.next_node] = [int(float(curr_node.distance))]
            curr_node = curr_node.next_node_reference


    def insert(self, trainLine, pNode, nNode, distance):
        if self.head is None:
            new_node = Node(distance if distance is not None else 0, next_node=nNode, prev_node=pNode, trainLine=trainLine)
            self.head = new_node
            return
        n = self.head
        while n.next_node_reference is not None:
            n = n.next_node_reference
        new_node = Node(distance if distance is not None else 0, next_node=nNode, prev_node=pNode, trainLine=trainLine)
        n.next_node_reference = new_node
        new_node.prev_node_reference = n
        self.end = n.next_node_reference


def readfile():
    workbook = openpyxl.load_workbook('London Underground data.xlsx')
    worksheet = workbook.active  # Assuming the first sheet is active
    trainLineList = []
    trainLineListCheck = []
    trainLineNum = -1

    for row in worksheet.iter_rows(values_only=True):
        if row[0] not in trainLineList:
            trainLineList.append(row[0])
            trainLineListCheck.append(row[0])

    trainLineList = [DoubleLinkedList() for _ in range(len(trainLineList))]

    for row in worksheet.iter_rows(values_only=True):
        if row[0] in trainLineListCheck:
            if row[0] == '':
                if worksheet.cell(row=row[1]+1, column=1).value == worksheet.cell(row=row[1]-1, column=1).value:
                    index = trainLineListCheck.index(worksheet.cell(row=row[1]-1, column=1).value)
                    trainLineList[index].insert(worksheet.cell(row=row[1]-1, column=1).value, row[1], row[2], row[3])
            else:
                if row[2] != '':
                    index = trainLineListCheck.index(row[0])
                    trainLineList[index].insert(row[0], row[1], row[2], row[3])
        else:
            print("error i believe there is no trainLine of the sort")

    return trainLineList



def dijkstra(graph, start, goal):
    shortest_distance = {}  # records the cost to reach to that node. Going to be
    track_predecessor = {}  # keep track of the path that has Led us to this node
    unseenNodes = graph  # to iterate through the entire graph.
    infinity = 9999999
    track_path = [] # going to trace our journey back to the source node -
    for node in unseenNodes:
        shortest_distance[node] = infinity
        shortest_distance[start] = 0

    while unseenNodes:

       min_distance_node = None
        
       for node in unseenNodes:
             if min_distance_node is None:
                 min_distance_node = node

             elif shortest_distance[node] < shortest_distance[min_distance_node]:
                 min_distance_node = node

       path_options = graph[min_distance_node].items()
       for child_node, weight in path_options:
           try:
                
               if weight[0] + shortest_distance[min_distance_node] < shortest_distance[child_node]:
                   shortest_distance[child_node] = weight[0] + shortest_distance[min_distance_node]
                   track_predecessor[child_node] = min_distance_node
           except KeyError:
               pass


       unseenNodes.pop(min_distance_node)

    currentNode = goal

    while currentNode != start:
        try:
            track_path.insert(0, currentNode)
            currentNode = track_predecessor[currentNode]

        except KeyError:
                print("Path-is-not.reachable")
                break

    track_path.insert(0,start)


    if shortest_distance[goal] != infinity:
               print("Shortest distance is " + str(shortest_distance[goal]))
               print("Optimal path is " + str(track_path))
               return shortest_distance[goal], track_path


class CreateGUI(Ui_MainWindow):
    def __init__(self, window):
        self.setupUi(window)
        self.findRouteButton.clicked.connect(self.calcRoute)
        self.tubeMapPopUpButton.clicked.connect(self.openTubeMap)

        font = QFont()
        font.setPointSize(15)
        self.ETA_edit.setFont(font)
        self.textEdit.setFont(font)

    def calcRoute(self):
        dt  = datetime.now()
        start = self.startStationEntry.text()
        destination = self.destinationStationEntry.text()
        timeTaken, path = dijkstra(graph, start, destination)
        if timeTaken is not None and path is not None:
            tempTime = dt + timedelta(minutes=timeTaken)
            self.textEdit.clear()
            self.textEdit.append
            result_text = "Optimal route:\n"
            for station in path:
                result_text+= station + "\n"
            journey_summary_text = f"Time taken is {timeTaken}\n"
            ETA_text = tempTime.strftime("%H:%M")
            self.textEdit.setText(result_text)
            self.ETA_edit.setText(ETA_text)
        

    def openTubeMap(self):
        tube_map_window = TubeMap()
        tube_map_window.show()

class TubeMap(QMainWindow):
    def __init__(self):
        super().__init__()




def main():
    global graph
    graph = {}
    trainLineList = readfile()
    for i in trainLineList:
        i.addToGraph(graph)
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = CreateGUI(MainWindow)
    MainWindow.show() 
    sys.exit(app.exec_())
    

if __name__ == "__main__":
    main()
